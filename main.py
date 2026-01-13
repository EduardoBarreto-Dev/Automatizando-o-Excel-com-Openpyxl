# -------------- Importações --------------
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment

'Font: Serve para alterar a fonte dos textos, mudar a cor e etc.'
'PatterFill: Serve para alterar o fundo das células.'
'Border: Serve para alterar a borda das células.'
'Side: Serve para definirt o tipo de borda das células'
'Alignment: Serve para definir o alinhamento dos textos das células'

# -------------- Istanciando as planilhas --------------
# Carregando a instância da planilha geral
alunos_gerais = load_workbook('alunos.xlsx')
alunos_planilha = alunos_gerais.active


# -------------- Configurações de Formatação --------------
def create_template(wb: Workbook, sheet_title: str, title: str, merge_cells: str, students: Workbook):
    '''
    Função responsável por criar um template padrão para as planilhas. Isso evita repetição de código e permite uma
    reutilização para várias planilhas que têm o mesmo propósito.

    :param wb: O Objeto que representa a planilha a ser modificada.
    :type planilha: Workbook

    :param sheet_title: O título da planilha ativa (a que está sendo modificada).
    :type sheet_title: str

    :param title: O título que ficará nas células mescladas.
    :type title: str

    :param merge_cells: As células que serão mescladas no formato "A1:D1" (exemplo).
    :type planilha: str

    Return
    ---
    `False` em caso de algum erro
    '''

    # -------------- Ativando a planilha --------------
    alunos = students.active
    planilha = wb.active

    # -------------- Pegando o cabeçalho da planilha alunos --------------
    'Adicionar Dados de um cabeçalho já existente evita reescrever o nome dos títulos dessas celulas'
    cabecalho_alunos = [celula.value for celula in alunos[1]]

    # -------------- Configurações de formatação --------------
    # Formatando o título
    config_font_title = Font(color = "ffffff", size = 14, bold = True)
    config_fill_title = PatternFill(fgColor = "747070", fill_type = 'solid')

    # Formatando o cabeçalho
    config_font_header = Font(color = "ffffff", size = 13, bold = True)
    config_fill_header = PatternFill(fgColor = "B4B0B0", fill_type = 'solid')

    # Formatação Geral
    config_align_title_and_header = Alignment(horizontal = 'center')
    thin_border = Side(style = 'thin')
    border = Border(top = thin_border, bottom = thin_border, left = thin_border, right = thin_border)

    # -------------- Aplicando formatação --------------
    try:
        # Juntando as células do cabeçalho
        planilha.merge_cells(merge_cells)

        # Inserindo o título da sheet
        planilha.title = sheet_title

        # Inserindo o título a linha 1
        planilha['A1'] = title

        # Pondo formatação do Título
        planilha['A1'].font = config_font_title
        planilha['A1'].fill = config_fill_title
        planilha['A1'].alignment = config_align_title_and_header
        planilha['A1'].border = border

        # Inserindo cabeçalho
        planilha.append(cabecalho_alunos)

        # Formatando o cabeçalho
        for celula in planilha[2]:
            celula.font = config_font_header
            celula.fill = config_fill_header
            celula.alignment = config_align_title_and_header
            celula.border = border

        # ------ Retornando o Objeto Workbook para salvar ------ 
        return wb


    except Exception as e:
        print (f'Ocorreu um erro: {e}')
        return False


# -------------- Criando e formatando as planilhas --------------
aprovados = create_template(Workbook(), 'Aprovados', "Alunos que estão Aprovados", 'A1:E1', alunos_gerais)
reprovados = create_template(Workbook(), 'Reprovados', "Alunos que estão Reprovados", 'A1:E1', alunos_gerais)

# Em caso de erro
if not aprovados or not reprovados:
    quit()

# Pegando a planilha de aprovados
aprovados_planilha = aprovados.active

# Pegando a planilha de reprovados
reprovados_planilha = reprovados.active


# -------------- Corpo principal --------------
alunos_aprovados = alunos_reprovados = maior_nota = 0
nome_maior_nota, media_turma = '', []


'Percorrer alunos os filtrando com base em suas notas'
for linha in alunos_planilha.iter_rows():
    'Cada linha é uma tupla de objetos célula. Cada objeto representa uma célula da planilha.'
    try:
        # Verificando a nota do aluno
        nota = linha[3].value
        
        # Ignora a primeira linha
        if nota == 'Nota Final':
            continue
        

        # Pegando as inforamções do aluno
        dados_aluno = [celula.value for celula in linha]
        dados_aluno[4] = dados_aluno[4].strftime('%d/%m/%Y') # Formatando a data

        # Dados para exibição posterior
        media_turma.append(nota)

        if nota > maior_nota:
            nome_maior_nota = dados_aluno[0]
            maior_nota = nota

        # Caso aluno esteja aprovado
        if nota >= 7:
            aprovados_planilha.append(dados_aluno)
            alunos_aprovados += 1
            continue
        
        # Caso esteja reprovado
        reprovados_planilha.append(dados_aluno)
        alunos_reprovados += 1
        
    # Caso ocorra algum erro
    except Exception as e:
        print(f"Ocorreu um erro: {e}")


# -------------- Exibindo informações de log --------------
print(f"Quantidade de alunos Aprovados: {alunos_aprovados}")
print(f"Quantidade de alunos Reprovados: {alunos_reprovados}")
print(f"Nota média da Turma: {sum(media_turma) / len(media_turma):.1f}")
print(f"Nota e nome do Aluno com maior nota: {nome_maior_nota} - {maior_nota}")



# -------------- Salvando as planilhas --------------
aprovados.save('aprovados.xlsx')

reprovados.save('reprovados.xlsx')

